"""
FastAPI backend for a contact form that saves each submission
as a new row in an Excel (.xlsx) file using openpyxl.

Install dependencies:
    pip install fastapi uvicorn openpyxl python-multipart

Run:
    uvicorn main:app --reload --port 8000

Endpoint:
    POST /api/contact
    Body (JSON): { "name": "...", "email": "...", "subject": "...", "message": "..." }
"""

import os
from datetime import datetime
from threading import Lock

from fastapi import FastAPI, HTTPException
from fastapi.middleware.cors import CORSMiddleware
from pydantic import BaseModel, EmailStr, Field
from openpyxl import Workbook, load_workbook

app = FastAPI(title="Contact Form API")

# Allow your Vue frontend (adjust origins as needed, e.g. your dev server / deployed domain)
app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],          # tighten this to your actual frontend URL in production
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

EXCEL_FILE = "contact_submissions.xlsx"
SHEET_NAME = "Submissions"
HEADERS = ["Timestamp", "Name", "Email", "Subject", "Message"]

# Excel writes aren't thread-safe; guard concurrent requests with a lock
file_lock = Lock()


class ContactForm(BaseModel):
    name: str = Field(..., min_length=1, max_length=100)
    email: EmailStr
    subject: str = Field(..., min_length=1, max_length=200)
    message: str = Field(..., min_length=1, max_length=2000)


def init_excel_file() -> None:
    """Create the Excel file with headers if it doesn't already exist."""
    if not os.path.exists(EXCEL_FILE):
        wb = Workbook()
        ws = wb.active
        ws.title = SHEET_NAME
        ws.append(HEADERS)
        for col_idx, header in enumerate(HEADERS, start=1):
            ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = 25
        wb.save(EXCEL_FILE)


def append_submission(data: ContactForm) -> None:
    """Append one row to the Excel sheet."""
    with file_lock:
        init_excel_file()
        wb = load_workbook(EXCEL_FILE)
        ws = wb[SHEET_NAME] if SHEET_NAME in wb.sheetnames else wb.active
        ws.append([
            datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            data.name,
            data.email,
            data.subject,
            data.message,
        ])
        wb.save(EXCEL_FILE)


@app.on_event("startup")
def on_startup():
    init_excel_file()


@app.post("/api/contact")
def submit_contact_form(form: ContactForm):
    try:
        append_submission(form)
    except Exception as exc:
        raise HTTPException(status_code=500, detail=f"Failed to save submission: {exc}")

    return {"status": "success", "message": "Thanks! Your message has been received."}


@app.get("/api/health")
def health_check():
    return {"status": "ok"}